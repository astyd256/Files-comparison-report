import os
from openpyxl import Workbook

def write_to_excel(file_list1, file_list2, output_file):
    wb = Workbook()
    ws = wb.active
    for i, file_name in enumerate(file_list1, start=1):
        if file_name in file_list2:
            ws.cell(row=i, column=1, value=file_name)
            ws.cell(row=i, column=2, value=file_name)
            file_list2.remove(file_name)
        else:
            ws.cell(row=i, column=1, value=file_name)
    for i, file_name in enumerate(file_list2, start=len(file_list1)+1):
        if ws.cell(row=i, column=2).value:
            pass
        else:
            ws.cell(row=i, column=2, value=file_name)
    wb.save(output_file)

folder_path1 = '/path/to/first/folder'
folder_path2 = '/path/to/second/folder'
output_file = 'excel_file_name.xlsx'

file_list1 = os.listdir(folder_path1)
file_list2 = os.listdir(folder_path2)
file_list1.sort()
file_list2.sort()
write_to_excel(file_list1, file_list2, output_file)
